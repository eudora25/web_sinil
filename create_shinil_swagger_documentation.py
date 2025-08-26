import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

def create_shinil_swagger_documentation_xlsx():
    # 신일 Swagger UI의 실제 API 데이터 정의
    api_data = [
        {
            "api_name": "거래원장 조회 API",
            "uri": "/getxxxxxxxxxxxx",
            "description": "거래원장 조회",
            "frequency": "실시간",
            "input_params": [
                {"name": "custcode", "type": "STRING", "description": "거래처코드", "required": "Y", "notes": ""},
                {"name": "sdt", "type": "STRING", "description": "조회 시작일", "required": "Y", "notes": ""},
                {"name": "edt", "type": "STRING", "description": "조회 종료일", "required": "Y", "notes": ""}
            ],
            "output_params": [
                {"name": "empname", "type": "STRING", "description": "영업사원명", "required": "", "notes": ""},
                {"name": "empname", "type": "STRING", "description": "영업사원부서명", "required": "", "notes": ""},
                {"name": "phone", "type": "STRING", "description": "연락처", "required": "", "notes": "암호화되어 있음"},
                {"name": "orderdate", "type": "STRING", "description": "거래일자", "required": "Y", "notes": ""},
                {"name": "itemname", "type": "STRING", "description": "제품명", "required": "Y", "notes": ""},
                {"name": "unit", "type": "STRING", "description": "단위", "required": "Y", "notes": ""},
                {"name": "salqty", "type": "FLOAT", "description": "수량", "required": "Y", "notes": ""},
                {"name": "salamt", "type": "FLOAT", "description": "매출액", "required": "Y", "notes": ""},
                {"name": "salvat", "type": "FLOAT", "description": "부가세", "required": "Y", "notes": ""},
                {"name": "colamt", "type": "FLOAT", "description": "수금액", "required": "Y", "notes": ""},
                {"name": "balace", "type": "FLOAT", "description": "잔액", "required": "Y", "notes": ""},
                {"name": "signyn", "type": "STRING", "description": "서명여부", "required": "Y", "notes": ""}
            ]
        },
        {
            "api_name": "사용자 등록 API",
            "uri": "/functions/v1/register-user",
            "description": "신규 사용자 등록",
            "frequency": "요청 시",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "notes": ""},
                {"name": "password", "type": "STRING", "description": "비밀번호", "required": "Y", "notes": ""},
                {"name": "company_name", "type": "STRING", "description": "회사명", "required": "Y", "notes": ""},
                {"name": "business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "notes": ""},
                {"name": "representative_name", "type": "STRING", "description": "대표자명", "required": "Y", "notes": ""},
                {"name": "business_address", "type": "STRING", "description": "사업장 주소", "required": "Y", "notes": ""},
                {"name": "contact_person_name", "type": "STRING", "description": "담당자명", "required": "Y", "notes": ""},
                {"name": "mobile_phone", "type": "STRING", "description": "연락처", "required": "Y", "notes": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "error", "type": "STRING", "description": "오류 메시지", "required": "N", "notes": "오류 발생 시"}
            ]
        },
        {
            "api_name": "비밀번호 재설정 이메일 전송 API",
            "uri": "/functions/v1/send-reset-email",
            "description": "비밀번호 재설정 이메일 전송",
            "frequency": "요청 시",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "notes": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "message", "type": "STRING", "description": "성공 메시지", "required": "Y", "notes": ""},
                {"name": "error", "type": "STRING", "description": "오류 메시지", "required": "N", "notes": "오류 발생 시"}
            ]
        },
        {
            "api_name": "비밀번호 초기화 API",
            "uri": "/functions/v1/reset-password",
            "description": "사용자 비밀번호 초기화",
            "frequency": "요청 시",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "notes": ""},
                {"name": "newPassword", "type": "STRING", "description": "새 비밀번호", "required": "Y", "notes": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "message", "type": "STRING", "description": "성공 메시지", "required": "Y", "notes": ""},
                {"name": "error", "type": "STRING", "description": "오류 메시지", "required": "N", "notes": "오류 발생 시"}
            ]
        },
        {
            "api_name": "사용자 생성 API (Vercel)",
            "uri": "/api/create-user",
            "description": "관리자 권한으로 사용자 생성",
            "frequency": "요청 시",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "notes": ""},
                {"name": "password", "type": "STRING", "description": "비밀번호", "required": "Y", "notes": ""},
                {"name": "company_name", "type": "STRING", "description": "회사명", "required": "N", "notes": ""}
            ],
            "output_params": [
                {"name": "user", "type": "OBJECT", "description": "생성된 사용자 정보", "required": "Y", "notes": ""},
                {"name": "error", "type": "STRING", "description": "오류 메시지", "required": "N", "notes": "오류 발생 시"}
            ]
        },
        {
            "api_name": "서버 상태 확인 API",
            "uri": "/api/health",
            "description": "백엔드 서버 상태 확인",
            "frequency": "요청 시",
            "input_params": [
                {"name": "없음", "type": "-", "description": "-", "required": "-", "notes": "-"}
            ],
            "output_params": [
                {"name": "status", "type": "STRING", "description": "서버 상태", "required": "Y", "notes": '"OK"'},
                {"name": "message", "type": "STRING", "description": "상태 메시지", "required": "Y", "notes": ""}
            ]
        },
        {
            "api_name": "실적 등록 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "실적 등록 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"},
                {"name": "company_id", "type": "INTEGER", "description": "업체 ID", "required": "Y", "notes": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "notes": "YYYY-MM 형식"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        },
        {
            "api_name": "도매 매출 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "도매 매출 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"},
                {"name": "company_id", "type": "INTEGER", "description": "업체 ID", "required": "Y", "notes": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "notes": "YYYY-MM 형식"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        },
        {
            "api_name": "직매 매출 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "직매 매출 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"},
                {"name": "company_id", "type": "INTEGER", "description": "업체 ID", "required": "Y", "notes": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "notes": "YYYY-MM 형식"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        },
        {
            "api_name": "증빙 파일 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "실적 증빙 파일 업로드",
            "frequency": "요청 시",
            "input_params": [
                {"name": "files", "type": "ARRAY[FILE]", "description": "증빙 파일들", "required": "Y", "notes": "최대 10개"},
                {"name": "company_id", "type": "INTEGER", "description": "업체 ID", "required": "Y", "notes": ""},
                {"name": "client_id", "type": "INTEGER", "description": "거래처 ID", "required": "Y", "notes": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "notes": "YYYY-MM 형식"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "uploaded_count", "type": "INTEGER", "description": "업로드된 파일 수", "required": "Y", "notes": ""},
                {"name": "file_paths", "type": "ARRAY", "description": "업로드된 파일 경로들", "required": "Y", "notes": ""}
            ]
        },
        {
            "api_name": "거래처-업체 매핑 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "거래처-업체 매핑 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        },
        {
            "api_name": "수수료 등급 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "수수료 등급 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        },
        {
            "api_name": "제품 정보 엑셀 업로드 API",
            "uri": "Supabase Storage + Database",
            "description": "제품 정보 엑셀 파일 업로드 및 처리",
            "frequency": "요청 시",
            "input_params": [
                {"name": "file", "type": "FILE", "description": "엑셀 파일", "required": "Y", "notes": ".xlsx, .xls"}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "notes": ""},
                {"name": "processed_count", "type": "INTEGER", "description": "처리된 건수", "required": "Y", "notes": ""},
                {"name": "error_count", "type": "INTEGER", "description": "오류 건수", "required": "Y", "notes": ""},
                {"name": "errors", "type": "ARRAY", "description": "오류 목록", "required": "N", "notes": ""}
            ]
        }
    ]

    # 워크북 생성
    wb = Workbook()
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 각 API에 대해 워크시트 생성
    for i, api in enumerate(api_data):
        if i == 0:
            ws = wb.active
            ws.title = f"{i+1}. {api['api_name']}"
        else:
            ws = wb.create_sheet(f"{i+1}. {api['api_name']}")
        
        # API 일반 정보
        ws['A1'] = "URI"
        ws['B1'] = api['uri']
        ws['A2'] = "설명"
        ws['B2'] = api['description']
        ws['A3'] = "주기"
        ws['B3'] = api['frequency']
        
        # 헤더 스타일 적용
        for cell in ['A1', 'A2', 'A3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].border = border
        
        # Input 파라미터 섹션
        ws['A5'] = "input"
        ws['A5'].font = section_font
        ws['A5'].fill = section_fill
        ws['A5'].border = border
        
        # Input 파라미터 헤더
        ws['A6'] = "파라미터"
        ws['B6'] = "타입"
        ws['C6'] = "설명"
        ws['D6'] = "필수구분"
        ws['E6'] = "비고"
        
        # Input 파라미터 헤더 스타일
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}6'].font = header_font
            ws[f'{col}6'].fill = header_fill
            ws[f'{col}6'].border = border
        
        # Input 파라미터 데이터
        for idx, param in enumerate(api['input_params']):
            row = 7 + idx
            ws[f'A{row}'] = param['name']
            ws[f'B{row}'] = param['type']
            ws[f'C{row}'] = param['description']
            ws[f'D{row}'] = param['required']
            ws[f'E{row}'] = param['notes']
            
            # 셀 스타일 적용
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = border
        
        # Output 파라미터 섹션
        output_start_row = 7 + len(api['input_params']) + 2
        ws[f'A{output_start_row}'] = "output"
        ws[f'A{output_start_row}'].font = section_font
        ws[f'A{output_start_row}'].fill = section_fill
        ws[f'A{output_start_row}'].border = border
        
        # Output 파라미터 헤더
        ws[f'A{output_start_row + 1}'] = "파라미터"
        ws[f'B{output_start_row + 1}'] = "타입"
        ws[f'C{output_start_row + 1}'] = "설명"
        ws[f'D{output_start_row + 1}'] = "필수구분"
        ws[f'E{output_start_row + 1}'] = "비고"
        
        # Output 파라미터 헤더 스타일
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{output_start_row + 1}'].font = header_font
            ws[f'{col}{output_start_row + 1}'].fill = header_fill
            ws[f'{col}{output_start_row + 1}'].border = border
        
        # Output 파라미터 데이터
        for idx, param in enumerate(api['output_params']):
            row = output_start_row + 2 + idx
            ws[f'A{row}'] = param['name']
            ws[f'B{row}'] = param['type']
            ws[f'C{row}'] = param['description']
            ws[f'D{row}'] = param['required']
            ws[f'E{row}'] = param['notes']
            
            # 셀 스타일 적용
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = border
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 25
    
    # 요약 시트 생성
    summary_ws = wb.create_sheet("API 요약")
    
    # 요약 헤더
    summary_headers = ["API명", "URI", "설명", "주기", "HTTP 메서드", "인증 필요", "Input 파라미터 수", "Output 파라미터 수", "비고"]
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # 요약 데이터
    summary_data = []
    for api in api_data:
        summary_data.append([
            api['api_name'],
            api['uri'],
            api['description'],
            api['frequency'],
            "POST" if "업로드" in api['api_name'] or "등록" in api['api_name'] or "생성" in api['api_name'] else "GET" if "상태" in api['api_name'] or "조회" in api['api_name'] else "POST",
            "필요" if "업로드" in api['api_name'] or "등록" in api['api_name'] or "생성" in api['api_name'] or "초기화" in api['api_name'] else "불필요",
            len(api['input_params']),
            len(api['output_params']),
            "엑셀 파일 처리" if "엑셀" in api['api_name'] else "파일 업로드" if "파일" in api['api_name'] else "사용자 관리" if "사용자" in api['api_name'] else "거래원장 조회" if "거래원장" in api['api_name'] else "시스템"
        ])
    
    for row_idx, row_data in enumerate(summary_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # 요약 시트 컬럼 너비 조정
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        summary_ws.column_dimensions[col].width = 15
    
    # 파일 저장
    wb.save('신일_Swagger_API_문서.xlsx')
    print("신일 Swagger API 문서가 성공적으로 생성되었습니다: 신일_Swagger_API_문서.xlsx")

if __name__ == "__main__":
    create_shinil_swagger_documentation_xlsx()





